#!/usr/bin/env python3
"""ETL: SpotsNow_Investor_Pipeline.xlsx -> vc-fund-data.js for the VC CRM.

- Excludes the Specialist (Parked) tab per Cam.
- Merges Cam's curated hand-grades (curated-seeds.json) onto matching funds.
- First-pass rubric grading from sheet signals; low-signal funds stay ungraded.
- Captures per-fund context (status, quotes, verdicts, actions) for the drawer.
- Builds connectors from the NFX Intro Paths tab (52 connectors, 2,239 edges).
- Seeds pipeline stages from live statuses (stage0 / archived0).
"""
import json, re, sys, unicodedata
import openpyxl

XLSX = '/Users/campbell/Downloads/SpotsNow_Investor_Pipeline.xlsx'
CURATED = 'curated-seeds.json'
OUT = '/Users/campbell/Desktop/Cursor Work/Newsletter page/vc-fund-data.js'

def slug(s):
    s = unicodedata.normalize('NFD', (s or '').lower())
    s = ''.join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r'[^a-z0-9]+', '-', s).strip('-')
    return s

STOP = {'ventures','venture','capital','partners','partner','fund','vc','the','group','company','llc','lp'}
def norm_org(s):
    s = re.sub(r'[^a-z0-9 ]', ' ', (s or '').lower())
    return ' '.join(w for w in s.split() if w and w not in STOP)

def clean(v):
    if v is None: return ''
    s = str(v).strip()
    return '' if s.lower() in ('none','nan','-') else s

wb = openpyxl.load_workbook(XLSX, data_only=True)
ws = wb['Investors']
head = [c.value for c in ws[1]]
ix = {h: i for i, h in enumerate(head)}
rows = list(ws.iter_rows(min_row=2, values_only=True))
def col(r, name):
    return clean(r[ix[name]]) if name in ix else ''

curated = {norm_org(f['name']): f for f in json.load(open(CURATED))}

# ── NFX intro paths: per-firm best edge + connector roster ──
ws2 = wb['NFX Intro Paths']
h2 = [c.value for c in ws2[1]]
i2 = {h: i for i, h in enumerate(h2)}
nfx_by_firm = {}   # norm firm -> list of rows
for r in ws2.iter_rows(min_row=2, values_only=True):
    firm = clean(r[i2['Firm']])
    if not firm: continue
    nfx_by_firm.setdefault(norm_org(firm), []).append({
        'strength': float(clean(r[i2['Intro Strength']]) or 0),
        'connector': clean(r[i2['Top Connector']]),
        'paths': clean(r[i2['Intro Paths']]),
        'name': clean(r[i2['Name']]),
        'title': clean(r[i2['Title']]),
        'check': clean(r[i2['Check Range']]),
        'sweet': clean(r[i2['Sweet Spot']]),
        'stages': clean(r[i2['Stages']]),
        'sectors': clean(r[i2['Matched Sectors']]),
        'geos': clean(r[i2['Geographies']]),
    })

# ── People tab (verified key people) ──
wsp = wb['People']
hp = [c.value for c in wsp[1]]
ip = {h: i for i, h in enumerate(hp)}
people_tab = {}
for r in wsp.iter_rows(min_row=2, values_only=True):
    firm = clean(r[ip['Firm']])
    if not firm: continue
    people_tab.setdefault(norm_org(firm), []).append({
        'n': clean(r[ip['Person']]), 't': clean(r[ip['Title']]) or 'Partner',
        'li': clean(r[ip['LinkedIn']]), 'why': clean(r[ip['Why They Matter']])
    })

# ── Signal keyword sets for rubric thesis scoring ──
SURFACE = {
    'marketplace': ['marketplace','marketplaces','network effects','two-sided'],
    'media': ['media','creator','audio','podcast','content','entertainment','consumer internet','community','passion economy','influencer'],
    'adtech': ['adtech','ad tech','advertising','martech','madtech'],
    'ai': ['ai','artificial intelligence','machine learning','data','ml'],
    'saas': ['saas','b2b software','vertical software','software'],
}
MISMATCH = ['bio','biotech','health','medical','pharma','therapeutic','fintech','climate','energy','crypto','web3','blockchain','hardware','deep tech','deeptech','defense','space','agriculture','real estate','proptech','insurance','life science','longevity','cannabis','mining','industrial']

def parse_check_k(txt):
    """'$500K - $1M' / '250000' style -> [min_k, max_k] or None"""
    if not txt: return None
    t = txt.lower().replace(',', '')
    nums = re.findall(r'\$?([\d.]+)\s*(m|k)?', t)
    vals = []
    for n, unit in nums:
        try: v = float(n)
        except ValueError: continue
        if unit == 'm': v *= 1000
        elif unit == 'k': pass
        elif v >= 10000: v = v / 1000       # raw dollars
        elif v < 20: v *= 1000              # bare millions
        vals.append(round(v))
    vals = [v for v in vals if v > 0][:2]
    if not vals: return None
    if len(vals) == 1: vals.append(vals[0])
    return [int(min(vals)), int(max(vals))]

def region_of(hq, geos):
    t = (hq + ' ' + geos).lower()
    if any(k in t for k in ['nashville','atlanta','chattanooga','tennessee','georgia','carolina','miami','florida','austin','texas']): return 'Southeast / South'
    if any(k in t for k in ['new york','nyc','brooklyn']): return 'New York'
    if any(k in t for k in ['san francisco','bay area','menlo','palo alto','sf']): return 'SF Bay Area'
    if any(k in t for k in ['los angeles','santa monica','la,']): return 'Los Angeles'
    if 'boston' in t or 'cambridge (ma)' in t: return 'Boston'
    if any(k in t for k in ['london','berlin','paris','europe','amsterdam','stockholm','netherlands','uk','germany','france']): return 'Europe'
    if any(k in t for k in ['united states','usa','remote']): return 'US (other)'
    if t.strip(): return 'International'
    return ''

funds = []
connectors = {}   # name -> {matches:[], notes:{}, strengths:[]}
seen_ids = set()
counts = {'graded': 0, 'curated': 0, 'ungraded': 0, 'people': 0, 'staged': 0, 'archived': 0}

for r in rows:
    name = col(r, 'Fund Name')
    if not name: continue
    nn = norm_org(name)
    fid = 'p-' + slug(name)
    while fid in seen_ids: fid += '-x'
    seen_ids.add(fid)

    tier_raw = col(r, 'Tier')
    tier_n = int(tier_raw.split(' ')[0]) if tier_raw[:1].isdigit() else 99
    nfx = sorted(nfx_by_firm.get(nn, []), key=lambda x: -x['strength'])
    best_nfx = nfx[0] if nfx else None

    # people
    ppl, pseen = [], set()
    def addp(n, t, li=''):
        n = clean(n);
        if not n or '@' in n or len(n) > 60: return
        k = slug(n)
        if not k or k in pseen: return
        pseen.add(k)
        ppl.append({'n': n, 't': clean(t) or 'Partner', **({'li': li} if li else {})})
    for pt in people_tab.get(nn, []): addp(pt['n'], pt['t'], pt['li'])
    pm = col(r, 'Partner / Fund Manager')
    if pm and len(pm) < 80:
        for part in re.split(r'\s*(?:/|&|,| and )\s*', pm)[:3]:
            if len(part.split()) in (2, 3): addp(part, 'Partner', col(r, 'Partner LinkedIn') if '/' not in pm else '')
    bc = col(r, 'Best Contact (researched)') or col(r, 'Who To Contact')
    if bc:
        m = re.match(r'^([A-Z][a-zA-Z.-]+ [A-Z][a-zA-Z.-]+)', bc)
        if m: addp(m.group(1), 'Contact')
    for e in nfx[:4]: addp(e['name'], e['title'] or 'Investor')
    if not ppl: ppl = [{'n': '', 't': 'Partner', 'tbd': True}]
    counts['people'] += sum(1 for p in ppl if not p.get('tbd'))

    # thesis signals
    sig_text = ' '.join([col(r, 'Focus Tags'), col(r, 'Focus (source)'),
                         best_nfx['sectors'] if best_nfx else '',
                         col(r, 'Description')[:600], col(r, 'Investor Type')]).lower()
    surfaces = {k for k, kws in SURFACE.items() if any(kw in sig_text for kw in kws)}
    mismatched = any(kw in sig_text for kw in MISMATCH)
    core = len(surfaces - {'saas'})
    has_signal = bool(sig_text.strip()) and (surfaces or mismatched or len(sig_text) > 80)

    # stage signals
    stages_txt = ((best_nfx['stages'] if best_nfx else '') + ' ' + col(r, 'Stages (source)') + ' ' + col(r, 'Earliest Stage')).lower()
    leads_seed = bool(col(r, 'Leads Seed')) or tier_raw.startswith('7')
    leads_pre = bool(col(r, 'Leads Pre-Seed'))
    early = ('pre-seed' in stages_txt) or ('seed' in stages_txt)
    late = any(k in stages_txt for k in ['series b', 'series c', 'growth']) and not early

    # check
    ck = None
    cm, cx = col(r, 'Check Min (USD)'), col(r, 'Check Max (USD)')
    if cm or cx:
        ck = parse_check_k((cm or cx) + ' - ' + (cx or cm))
    if not ck and best_nfx: ck = parse_check_k(best_nfx['check'])
    if not ck: ck = parse_check_k(col(r, 'Check Size (researched)') or col(r, 'Check Size (source)'))

    hq = col(r, 'HQ / Location')
    region = region_of(hq, best_nfx['geos'] if best_nfx else '') or (hq[:24] if hq else '')

    # ── dims (rubric first pass) ──
    dims = None
    if nn in curated:
        cf = curated[nn]
        dims = cf.get('dims')
        counts['curated'] += 1
    elif has_signal:
        if mismatched and core == 0: thesis = 30
        elif core >= 3: thesis = 90
        elif core >= 2: thesis = 84
        elif 'marketplace' in surfaces: thesis = 78
        elif 'media' in surfaces or 'adtech' in surfaces: thesis = 74
        elif 'ai' in surfaces: thesis = 66
        elif surfaces: thesis = 60
        else: thesis = 52
        if mismatched and core > 0: thesis = min(thesis, 64)
        stage = 88 if (leads_seed or leads_pre) else 80 if early else 45 if late else 64
        if ck:
            lo, hi = ck
            check = 88 if (lo <= 1500 and hi >= 250) else 62 if hi < 150 else 55 if lo > 3000 else 74
        else: check = 68
        portfolio = 80 if (best_nfx and 'marketplace' in (best_nfx['sectors'] or '').lower()) else 68 if core >= 1 else 58
        geo = {'Southeast / South': 92, 'New York': 76, 'SF Bay Area': 72, 'Los Angeles': 70,
               'Boston': 68, 'US (other)': 66, 'Europe': 50, 'International': 42}.get(region, 60)
        if best_nfx and best_nfx['strength'] >= 8: geo = max(geo, 82)
        elif best_nfx and best_nfx['strength'] >= 5: geo = max(geo, 74)
        dims = {'thesis': thesis, 'stage': stage, 'check': check, 'portfolio': portfolio, 'geo': geo}
        counts['graded'] += 1
    else:
        counts['ungraded'] += 1

    # ── stage0 / archived0 from live pipeline ──
    status = col(r, 'Status')
    crm = col(r, 'CRM Deal Status')
    stage0 = None
    archived0 = False
    if crm in ('Term Sheet', 'Diligence'): stage0 = 'pitching'
    elif status in ('Scheduled', 'Engaged 🔥'): stage0 = 'booked'
    elif tier_n == 1: stage0 = 'pitching'
    elif tier_n == 2 or crm in ('Targets 🎯', 'To Send', 'Waiting', 'Try Recover') or status in ('Todo', 'Get Intro', 'Waiting', 'Contacted'): stage0 = 'target'
    # Passes and not-relevant funds are pipeline history, not trash: they land
    # in the Closed stage, visible, never pre-archived (per Cam).
    if tier_n == 9 or status in ('Not Relevant', 'Pass / Update list') or ('Passed' in status and tier_n not in (1, 2, 3)) or 'Competitive' in status:
        stage0 = 'closed'
    if stage0: counts['staged'] += 1
    if archived0: counts['archived'] += 1

    # ── context for the drawer ──
    ctx = []
    def addctx(label, *cols_):
        vals = [col(r, c) for c in cols_]
        v = ' · '.join(x for x in vals if x)
        if v: ctx.append([label, v[:600]])
    addctx('Status', 'Status', 'CRM Deal Status', 'Tier')
    addctx('Last contact', 'Last Contact', 'Last Touch', 'Months Since')
    addctx('What happened', 'What Happened')
    addctx('They said', 'Verbatim Quote', 'What They Said')
    addctx('Where it was left', 'Where It Was Left')
    addctx('Action needed', 'ACTION NEEDED', 'Next Action')
    addctx('Re-approach', 'Re-approach Verdict', 'Verdict Confidence', 'Re-approach Angle')
    addctx('Pass reason', 'Pass Reason', 'Structural or Fixable', 'Disqualifier')
    # structured intro paths (synthesized 'ways in')
    # Guard: sheet warm-path columns sometimes hold the TARGET's own LinkedIn
    # (the partner at the fund), which is a contact, not a relationship. Detect
    # that, attach the URL to the person instead, and drop the bogus path.
    paths = []
    def _person_key_of(via):
        m = re.search(r'linkedin\.com/in/([^/?#]+)', via, re.I)
        cand = slug(m.group(1)) if m else slug(via)
        cand = cand.replace('-', '')
        if not cand: return None
        for k in pseen:
            kk = k.replace('-', '')
            if kk and (kk == cand or kk in cand or cand in kk):
                return k
        return None
    def add_path(via, kind, note):
        pk = _person_key_of(via)
        if pk:
            m = re.search(r'https?://\S+', via)
            if m:
                for p0 in ppl:
                    if not p0.get('tbd') and slug(p0['n']) == pk and not p0.get('li'):
                        p0['li'] = m.group(0).rstrip('|,; ')
            return
        paths.append({'via': via[:120], 'kind': kind, 'note': note})
    wip = col(r, 'Warm Intro Paths')
    if wip: add_path(wip, 'warm', 'listed as a warm intro path in the sheet')
    intro_from = col(r, 'Intro From (existing)')
    if intro_from: add_path(intro_from, 'existing', 'has introduced us before')
    crm_conn = col(r, 'CRM Connector')
    if crm_conn: add_path(crm_conn, 'existing', 'connector in Cams CRM')
    nfx_conn_col = col(r, 'NFX Connector')
    nfx_strength_col = col(r, 'NFX Intro Strength')
    if nfx_conn_col and not nfx:
        note = ('NFX strength ' + nfx_strength_col.rstrip('.0')) if nfx_strength_col else 'NFX Signal'
        paths.append({'via': nfx_conn_col[:120], 'kind': 'nfx', 'note': note})
    addctx('Pitched as', 'Pitched As (era)', 'Email Round')
    addctx('New fund', 'New Fund 2025-26', 'New Fund Detail')
    addctx('Sheet notes', 'Notes')
    nl = col(r, 'Notion Link')
    if nl: ctx.append(['Notion', nl])

    desc = col(r, 'Description')
    desc = re.sub(r'^#+\s*', '', desc.replace('\n', ' ')).strip()[:420]

    f = {
        'id': fid, 'name': name, 'site': col(r, 'Domain') or re.sub(r'^https?://(www\.)?', '', col(r, 'Website')).rstrip('/'),
        'type': col(r, 'Investor Type') or ('Seed' if (leads_seed or leads_pre or early) else '—'),
        'region': region or '—',
        'check': ck, 'sectors': [t.strip() for t in col(r, 'Focus Tags').split(';') if t.strip()][:3],
        'tier': tier_raw, 'people': ppl,
        'looking': desc, 'ctx': ctx,
    }
    if dims: f['dims'] = dims
    if nn in curated:
        cf = curated[nn]
        for k in ('looking', 'hook', 'why', 'sectors'):
            if cf.get(k): f[k] = cf[k]
        if not f.get('check') and cf.get('check'): f['check'] = cf['check']
        if cf.get('region') and f['region'] in ('—', ''): f['region'] = cf['region']
        if cf.get('type'): f['type'] = cf['type']
    if paths: f['paths'] = paths
    if stage0: f['stage0'] = stage0
    if archived0: f['archived0'] = True
    f['_nn'] = nn; f['_slug'] = slug(name); f['_tier_n'] = tier_n
    funds.append(f)

    # ── connector edges from every NFX path for this firm ──
    for e in nfx:
        cn = e['connector']
        if not cn: continue
        c = connectors.setdefault(cn, {'matches': [], 'notes': {}, 'strengths': []})
        target_pid = None
        if e['name']:
            k = slug(e['name'])
            if k in pseen: target_pid = fid + '_' + k
        if not target_pid:
            first = ppl[0]
            target_pid = fid + '_' + ('tbd-' + slug(first['t']) if first.get('tbd') else slug(first['n']))
        if target_pid not in c['matches']:
            c['matches'].append(target_pid)
            note = f"NFX strength {int(e['strength'])}" + (f" · {int(float(e['paths']))} paths" if e['paths'] else '')
            c['notes'][target_pid] = note
            c['strengths'].append(e['strength'])

# ── Dedupe pass: same normalized name + slug containment => same fund ──
drop_map = {}
survivors = []
by_nn = {}
for f in funds:
    nn = f['_nn']
    merged = False
    for surv in by_nn.get(nn, []):
        a, b = surv['_slug'], f['_slug']
        if a.startswith(b) or b.startswith(a):
            # merge f into surv
            drop_map[f['id']] = surv['id']
            for k in ('site','looking'):
                if not surv.get(k) and f.get(k): surv[k] = f[k]
            for k in ('type','region'):
                if surv.get(k) in ('', '—') and f.get(k) not in ('', '—'): surv[k] = f[k]
            if not surv.get('check') and f.get('check'): surv['check'] = f['check']
            if not surv.get('dims') and f.get('dims'): surv['dims'] = f['dims']
            for kk in ('hook','why'):
                if not surv.get(kk) and f.get(kk): surv[kk] = f[kk]
            if f['_tier_n'] < surv['_tier_n']:
                surv['tier'] = f['tier']; surv['_tier_n'] = f['_tier_n']
            if f.get('stage0') and not surv.get('stage0'): surv['stage0'] = f['stage0']
            if surv.get('stage0'): surv.pop('archived0', None)
            elif f.get('archived0'): surv['archived0'] = True
            pv = {(x['via'], x['kind']) for x in surv.get('paths', [])}
            for x in f.get('paths', []):
                if (x['via'], x['kind']) not in pv:
                    surv.setdefault('paths', []).append(x); pv.add((x['via'], x['kind']))
            sec = list(dict.fromkeys((surv.get('sectors') or []) + (f.get('sectors') or [])))[:3]
            surv['sectors'] = sec
            have = {slug(p0['n']) for p0 in surv['people'] if not p0.get('tbd')}
            for p0 in f['people']:
                if p0.get('tbd'): continue
                if slug(p0['n']) not in have:
                    surv['people'].append(p0); have.add(slug(p0['n']))
            if any(not p0.get('tbd') for p0 in surv['people']):
                surv['people'] = [p0 for p0 in surv['people'] if not p0.get('tbd')]
            seen_ctx = {tuple(c) for c in surv['ctx']}
            for c in f['ctx']:
                if tuple(c) not in seen_ctx: surv['ctx'].append(c); seen_ctx.add(tuple(c))
            merged = True
            break
    if not merged:
        by_nn.setdefault(nn, []).append(f)
        survivors.append(f)
funds = survivors
for f in funds:
    f.pop('_nn', None); f.pop('_slug', None); f.pop('_tier_n', None)

# remap connector edges from dropped fund ids to survivors
def remap_pid(pid):
    fid, _, rest = pid.partition('_')
    return (drop_map.get(fid, fid)) + '_' + rest
for c in connectors.values():
    new_matches, new_notes = [], {}
    for pid in c['matches']:
        np_ = remap_pid(pid)
        if np_ not in new_matches:
            new_matches.append(np_)
            new_notes[np_] = c['notes'].get(pid, '')
    c['matches'], c['notes'] = new_matches, new_notes
print('deduped:', len(drop_map), 'rows merged ->', len(funds), 'unique funds')

conn_out = []
for cname, c in sorted(connectors.items(), key=lambda kv: -len(kv[1]['matches'])):
    avg = sum(c['strengths']) / max(1, len(c['strengths']))
    conn_out.append({
        'id': 'c_nfx_' + slug(cname), 'name': cname,
        'role': 'NFX Signal connector',
        'rapport': f"Mapped from NFX Signal: {len(c['matches'])} intro paths, avg strength {avg:.1f}/10. Confirm the real-world rapport before asking.",
        'li': '', 'matches': c['matches'], 'matchNotes': c['notes'], 'source': 'nfx'
    })

out = {'generated': '2026-08-23', 'source': 'SpotsNow_Investor_Pipeline.xlsx (Specialist tab excluded)',
       'funds': funds, 'connectors': conn_out}
js = 'window.SN_DATA = ' + json.dumps(out, ensure_ascii=False, separators=(',', ':')) + ';\n'
open(OUT, 'w').write(js)
print('funds:', len(funds), counts)
print('connectors:', len(conn_out), 'edges:', sum(len(c['matches']) for c in conn_out))
print('output size: %.1f MB' % (len(js) / 1e6))
